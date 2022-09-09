import tushare as ts
import datetime
import base
import sys
import openpyxl

wb = openpyxl.load_workbook('earn.xlsx')
#获取工作表名
# sheet = wb.sheetnames
ws = wb['earn']
rows = []
for row in ws.rows:
    rows.append(row)
new_line = len(rows) + 1

today = sys.argv[1] if len(sys.argv) > 1 else datetime.datetime.now().strftime("%Y-%m-%d")
# today = "2019-06-03"
stock = []
code = ('600606', '600549', '000159', '300519', '002930', '600929', '300505', '601138', '002119', '300099', '002847', '600581', '603259', '300265', '600290', '600352', '300644', '600128', '601860', '600093', '002905', '300779', '603183', '002517', '300615', '603797')
ws.cell(row=new_line, column=1).value = today
n = 2
for i in code:
    data = ts.get_hist_data(i, today, today)
    gap_ma5 = 0
    gap_ma10 = 0
    gap_ma20 = 0
    if not data.empty:
        daily = data.iloc[0]
        gap_ma5 = round((daily.loc['close']/daily.loc['ma5']-1)*100, 2)
        gap_ma10 = round((daily.loc['close']/daily.loc['ma10']-1)*100, 2)
        gap_ma20 = round((daily.loc['close']/daily.loc['ma20']-1)*100, 2)
    ws.cell(row=new_line, column=n).value = gap_ma5
    n += 1
    ws.cell(row=new_line, column=n).value = gap_ma10
    n += 1
    ws.cell(row=new_line, column=n).value = gap_ma20
    n += 1
    #print((stockInfo['name'], gap_ma5, gap_ma10, gap_ma20))

wb.save('earn.xlsx') # 名称重复会覆盖
sys.exit()
# stock.sort(key=sort_stock)
#print(stock)
