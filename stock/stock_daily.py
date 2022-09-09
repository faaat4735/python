import tushare as ts
import datetime
import base
import sys
import openpyxl

wb = openpyxl.load_workbook('earn.xlsx')
#获取工作表名
# sheet = wb.sheetnames
ws = wb['earn']
rows = []
for row in ws.rows:
    rows.append(row)
new_line = len(rows) + 1

end = datetime.datetime.now()
if len(sys.argv) > 1:
    input_d = sys.argv[1].split('-')
    input_d = list(map(int, input_d))
    begin = datetime.datetime(input_d[0], input_d[1], input_d[2])
else :
    begin = end

# today = "2019-06-03"
stock = []
code = ('600606', '600549', '000159', '300519', '002930', '600929', '300505', '601138', '002119', '300099', '002847', '600581', '603259', '300265', '600290', '600352', '300644', '600128', '601860', '600093', '002905', '300779', '603183', '002517', '300615')


delta = datetime.timedelta(days=1)
while begin <= end:
    today = begin.strftime("%Y-%m-%d")
    data = ts.get_hist_data('600606', today, today)
    begin += delta
    if data.empty:
        continue
    print(today)
    ws.cell(row=new_line, column=1).value = today
    n = 2
    for i in code:
        data = ts.get_hist_data(i, today, today)
        daily = data.iloc[0]
        gap_ma5 = round((daily.loc['close']/daily.loc['ma5']-1)*100, 2)
        gap_ma10 = round((daily.loc['close']/daily.loc['ma10']-1)*100, 2)
        gap_ma20 = round((daily.loc['close']/daily.loc['ma20']-1)*100, 2)
        ws.cell(row=new_line, column=n).value = gap_ma5
        n += 1
        ws.cell(row=new_line, column=n).value = gap_ma10
        n += 1
        ws.cell(row=new_line, column=n).value = gap_ma20
        n += 1
        #print((stockInfo['name'], gap_ma5, gap_ma10, gap_ma20))
    new_line += 1

wb.save('earn.xlsx') # 名称重复会覆盖
sys.exit()
# stock.sort(key=sort_stock)
#print(stock)
